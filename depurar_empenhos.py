import csv
import logging
import os
import unicodedata

import openpyxl

from config import WORKBOOK_PATH, REPORT_FILE

SHEET_CORH = "ajustada CORH"
SHEET_BALANCES = "SALDOS DE EMPENHO"
BALANCES_START_ROW = 7

DETAILED = True
LOG_LEVEL = "INFO"

logging.basicConfig(
    level=getattr(logging, LOG_LEVEL),
    format="%(levelname)-7s %(message)s",
)
log = logging.getLogger("debugger")


def normalize_text(text):
    result = str(text or "").upper().strip()
    result = "".join(
        c for c in unicodedata.normalize("NFD", result)
        if unicodedata.category(c) != "Mn"
    )
    result = result.replace("-", " ")
    while "  " in result:
        result = result.replace("  ", " ")
    return result.strip()


def extract_commitment_year(note):
    year = 2020
    if len(note) >= 15:
        year_str = note[11:15]
        if year_str.isdigit():
            n = int(year_str)
            if 2000 <= n <= 2100:
                year = n
    return year


def to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def as_str(v):
    return str(v).strip() if v is not None else ""


def last_row(ws, col=1):
    lr = ws.max_row
    while lr > 1 and ws.cell(lr, col).value in (None, ""):
        lr -= 1
    return lr


def money(v):
    return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def debug(path=WORKBOOK_PATH):
    if not os.path.exists(path):
        log.error("Arquivo nao encontrado: %s", path)
        return

    log.info("=" * 60)
    log.info(" DEPURADOR - PROCESSAR EMPENHOS (dry-run, sem alterar o arquivo)")
    log.info("=" * 60)
    log.info("Arquivo: %s", os.path.abspath(path))

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    if SHEET_CORH not in wb.sheetnames:
        log.error("Aba '%s' nao encontrada. Abas disponiveis: %s", SHEET_CORH, wb.sheetnames)
        return
    if SHEET_BALANCES not in wb.sheetnames:
        log.error("Aba '%s' nao encontrada. Abas disponiveis: %s", SHEET_BALANCES, wb.sheetnames)
        return

    ws_corh = wb[SHEET_CORH]
    ws_balances = wb[SHEET_BALANCES]

    last_row_corh = last_row(ws_corh)
    last_row_balances = last_row(ws_balances)
    log.info("Aba CORH   '%s' -> ultima linha %d", SHEET_CORH, last_row_corh)
    log.info("Aba SALDOS '%s' -> ultima linha %d", SHEET_BALANCES, last_row_balances)

    notes, sources, balances, years, counts = {}, {}, {}, {}, {}
    stats = {
        "read": 0, "empty_contract": 0, "empty_note": 0,
        "zero_balance": 0, "invalid_year": 0, "valid": 0,
    }

    for i in range(BALANCES_START_ROW, last_row_balances + 1):
        stats["read"] += 1
        contract = as_str(ws_balances.cell(i, 1).value)
        if contract == "":
            stats["empty_contract"] += 1
            continue
        municipality_norm = normalize_text(as_str(ws_balances.cell(i, 13).value))
        note = as_str(ws_balances.cell(i, 9).value)
        if note == "":
            stats["empty_note"] += 1
            continue
        source = as_str(ws_balances.cell(i, 7).value)

        balance_o = to_float(ws_balances.cell(i, 15).value)
        balance_q = to_float(ws_balances.cell(i, 17).value)
        final_balance = balance_q if balance_q > 0 else balance_o
        if final_balance <= 0:
            stats["zero_balance"] += 1
            continue

        year = extract_commitment_year(note)
        if year < 2025 or year > 2026:
            stats["invalid_year"] += 1
            log.debug("Linha saldo %d: ano %d fora de 2025-2026 (nota %s)", i, year, note)
            continue

        key = contract + "_" + municipality_norm
        count = counts.get(key, 0) + 1
        counts[key] = count
        ck = key + "_" + str(count)
        notes[ck] = note
        sources[ck] = source
        balances[ck] = final_balance
        years[ck] = year
        stats["valid"] += 1

    log.info("-" * 60)
    log.info("CARREGAMENTO DE SALDOS")
    log.info("  Linhas lidas .............. %d", stats["read"])
    log.info("  Ignoradas (contrato vazio)  %d", stats["empty_contract"])
    log.info("  Ignoradas (nota vazia) .... %d", stats["empty_note"])
    log.info("  Ignoradas (saldo <= 0) .... %d", stats["zero_balance"])
    log.info("  Ignoradas (ano != 2025/26)  %d", stats["invalid_year"])
    log.info("  Empenhos validos .......... %d", stats["valid"])
    log.info("  Chaves distintas .......... %d", len(counts))

    corh_vals = {}
    for row in ws_corh.iter_rows(min_row=1, max_row=last_row_corh, min_col=1, max_col=13):
        for cell in row:
            if cell.value is not None:
                corh_vals[(cell.row, cell.column)] = cell.value

    def cval(r, c):
        return corh_vals.get((r, c))

    used_balances = {}
    problems = []
    summary = {"analyzed": 0, "ok": 0, "not_found": 0,
               "insufficient_balance": 0, "already_processed": 0, "no_amount": 0}

    log.info("-" * 60)
    log.info("ANALISE DAS LINHAS CORH (previsao)")

    for i in range(last_row_corh, 1, -1):
        contract = as_str(cval(i, 1))
        if contract == "":
            continue
        if as_str(cval(i, 8)) != "":
            summary["already_processed"] += 1
            if DETAILED:
                log.info("[linha %d] JA PROCESSADA (col H preenchida) -> seria pulada", i)
            continue

        municipality = as_str(cval(i, 3))
        municipality_norm = normalize_text(municipality)
        total_amount = to_float(cval(i, 13))
        if total_amount <= 0:
            summary["no_amount"] += 1
            if DETAILED:
                log.info("[linha %d] valor <= 0 -> seria pulada", i)
            continue

        summary["analyzed"] += 1
        key = contract + "_" + municipality_norm

        if key not in counts:
            summary["not_found"] += 1
            log.warning("[linha %d] PROBLEMA: contrato/municipio nao encontrado "
                        "(%s | %s)", i, contract, municipality)
            problems.append([i, contract, municipality, money(total_amount),
                             "Contrato/Municipio nao encontrado"])
            continue

        count = counts[key]
        commitments = []
        for j in range(1, count + 1):
            ck = key + "_" + str(j)
            commitments.append({"note": notes[ck], "balance": balances[ck],
                                "year": years[ck]})
        commitments.sort(key=lambda e: e["year"])

        remaining = total_amount
        used = 0
        usage_detail = []
        for e in commitments:
            if remaining <= 0:
                break
            current_balance = used_balances.get(e["note"], e["balance"])
            if current_balance <= 0:
                continue
            amount_to_use = remaining if remaining <= current_balance else current_balance
            used_balances[e["note"]] = current_balance - amount_to_use
            remaining -= amount_to_use
            used += 1
            usage_detail.append(f"{e['note']}={money(amount_to_use)}")

        if used == 0:
            summary["insufficient_balance"] += 1
            log.warning("[linha %d] PROBLEMA: sem saldo disponivel nos empenhos "
                        "(%s | %s) faltam %s", i, contract, municipality, money(remaining))
            problems.append([i, contract, municipality, money(remaining),
                             "Saldo insuficiente"])
        elif remaining > 0.01:
            summary["insufficient_balance"] += 1
            log.warning("[linha %d] PROBLEMA: saldo insuficiente, faltam %s "
                        "(%s | %s) | usaria %d empenho(s): %s",
                        i, money(remaining), contract, municipality,
                        used, "; ".join(usage_detail))
            problems.append([i, contract, municipality, money(remaining),
                             "Saldo insuficiente"])
        else:
            summary["ok"] += 1
            extra_rows = used - 1
            if DETAILED:
                log.info("[linha %d] OK: valor %s coberto por %d empenho(s)%s -> %s",
                         i, money(total_amount), used,
                         f" (+{extra_rows} linha[s])" if extra_rows else "",
                         "; ".join(usage_detail))

    log.info("-" * 60)
    log.info("RESUMO DA PREVISAO")
    log.info("  Linhas a processar ........ %d", summary["analyzed"])
    log.info("  OK ........................ %d", summary["ok"])
    log.info("  Nao encontrados ........... %d", summary["not_found"])
    log.info("  Saldo insuficiente ........ %d", summary["insufficient_balance"])
    log.info("  Ja processadas (puladas) .. %d", summary["already_processed"])
    log.info("  Sem valor (puladas) ....... %d", summary["no_amount"])

    dest = REPORT_FILE
    with open(dest, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["Linha", "Contrato", "Municipio", "Valor faltante", "Motivo"])
        w.writerows(problems)
    log.info("-" * 60)
    log.info("Relatorio de problemas salvo em: %s (%d problema[s])", dest, len(problems))


if __name__ == "__main__":
    debug()
