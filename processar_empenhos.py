import time
import unicodedata
from copy import copy

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from config import WORKBOOK_PATH

SHEET_CORH = "ajustada CORH"
SHEET_BALANCES = "SALDOS DE EMPENHO"
BALANCES_START_ROW = 7


def normalize_text(text):
    result = str(text or "").upper().strip()
    result = "".join(
        c for c in unicodedata.normalize("NFD", result)
        if unicodedata.category(c) != "Mn"
    )
    result = result.replace("-", " ")
    while "  " in result:
        result = result.replace("  ", " ")
    return result.strip()


def extract_commitment_year(note):
    year = 2020
    if len(note) >= 15:
        year_str = note[11:15]
        if year_str.isdigit():
            n = int(year_str)
            if 2000 <= n <= 2100:
                year = n
    return year


def to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def as_str(v):
    return str(v).strip() if v is not None else ""


def last_row(ws, col=1):
    lr = ws.max_row
    while lr > 1 and ws.cell(lr, col).value in (None, ""):
        lr -= 1
    return lr


def copy_row(ws, source, dest, max_col):
    for col in range(1, max_col + 1):
        c_src = ws.cell(row=source, column=col)
        c_dst = ws.cell(row=dest, column=col)
        c_dst.value = c_src.value
        if c_src.has_style:
            c_dst.font = copy(c_src.font)
            c_dst.border = copy(c_src.border)
            c_dst.fill = copy(c_src.fill)
            c_dst.alignment = copy(c_src.alignment)
            c_dst.number_format = c_src.number_format
            c_dst.protection = copy(c_src.protection)


def autofit(ws, first_col, last_col):
    for col in range(first_col, last_col + 1):
        letter = openpyxl.utils.get_column_letter(col)
        widest = 0
        for cell in ws[letter]:
            if cell.value is not None:
                widest = max(widest, len(str(cell.value)))
        ws.column_dimensions[letter].width = widest + 2


def create_log_sheet(wb):
    if "LOG" in wb.sheetnames:
        wb.remove(wb["LOG"])
    ws = wb.create_sheet("LOG")
    ws.cell(1, 1, "Linha")
    ws.cell(1, 2, "Contrato")
    ws.cell(1, 3, "Municipio")
    ws.cell(1, 4, "Valor faltante")
    ws.cell(1, 5, "Motivo")
    gray = PatternFill(start_color="C8C8C8", end_color="C8C8C8", fill_type="solid")
    for col in range(1, 6):
        c = ws.cell(1, col)
        c.font = Font(bold=True)
        c.fill = gray
        c.alignment = Alignment(horizontal="center", vertical="center")
    return ws


def process_commitments(path=WORKBOOK_PATH):
    start_time = time.time()

    wb = openpyxl.load_workbook(path, keep_vba=str(path).lower().endswith(".xlsm"))
    ws_corh = wb[SHEET_CORH]
    ws_balances = wb[SHEET_BALANCES]

    notes, sources, balances = {}, {}, {}
    years, counts, used_balances = {}, {}, {}

    last_row_corh = last_row(ws_corh)
    last_row_balances = last_row(ws_balances)

    if not ws_corh.cell(5, 15).value:
        ws_corh.cell(5, 15, "Saldo utilizado NE")
    if not ws_corh.cell(5, 16).value:
        ws_corh.cell(5, 16, "Saldo remanescente NE")
    if not ws_corh.cell(5, 17).value:
        ws_corh.cell(5, 17, "Controle Macro")

    ws_log = create_log_sheet(wb)
    log_row = 2

    for i in range(BALANCES_START_ROW, last_row_balances + 1):
        contract = as_str(ws_balances.cell(i, 1).value)
        if contract == "":
            continue
        municipality_norm = normalize_text(as_str(ws_balances.cell(i, 13).value))
        note = as_str(ws_balances.cell(i, 9).value)
        if note == "":
            continue
        source = as_str(ws_balances.cell(i, 7).value)

        balance_o = to_float(ws_balances.cell(i, 15).value)
        balance_q = to_float(ws_balances.cell(i, 17).value)
        final_balance = balance_q if balance_q > 0 else balance_o
        if final_balance <= 0:
            continue

        year = extract_commitment_year(note)
        if year < 2025 or year > 2026:
            continue

        key = contract + "_" + municipality_norm
        count = counts.get(key, 0) + 1
        counts[key] = count

        ck = key + "_" + str(count)
        notes[ck] = note
        sources[ck] = source
        balances[ck] = final_balance
        years[ck] = year

    inserted_rows = used_commitments = not_found = 0
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    orange = PatternFill(start_color="FFC864", end_color="FFC864", fill_type="solid")

    i = last_row_corh
    while i >= 2:
        contract = as_str(ws_corh.cell(i, 1).value)
        if contract == "":
            i -= 1
            continue
        if as_str(ws_corh.cell(i, 8).value) != "":
            i -= 1
            continue

        municipality = as_str(ws_corh.cell(i, 3).value)
        municipality_norm = normalize_text(municipality)
        total_amount = to_float(ws_corh.cell(i, 13).value)
        if total_amount <= 0:
            i -= 1
            continue

        key = contract + "_" + municipality_norm

        if key not in counts:
            for col in range(1, 17):
                ws_corh.cell(i, col).fill = yellow
            ws_log.cell(log_row, 1, i)
            ws_log.cell(log_row, 2, contract)
            ws_log.cell(log_row, 3, municipality)
            ws_log.cell(log_row, 4, total_amount)
            ws_log.cell(log_row, 5, "Contrato/Municipio nao encontrado")
            log_row += 1
            not_found += 1
            i -= 1
            continue

        count = counts[key]
        commitments = []
        for j in range(1, count + 1):
            ck = key + "_" + str(j)
            commitments.append({
                "note": notes[ck], "source": sources[ck],
                "balance": balances[ck], "year": years[ck],
            })
        commitments.sort(key=lambda e: e["year"])

        remaining = total_amount
        found = False

        for e in commitments:
            if remaining <= 0:
                break
            note, source, year = e["note"], e["source"], e["year"]
            current_balance = used_balances.get(note, e["balance"])
            if current_balance <= 0:
                continue

            amount_to_use = remaining if remaining <= current_balance else current_balance
            competence = 3 if year >= 2026 else 2

            if found:
                ws_corh.insert_rows(i + 1)
                copy_row(ws_corh, i, i + 1, 17)
                inserted_rows += 1
                ws_corh.cell(i + 1, 8, note[11:])
                ws_corh.cell(i + 1, 10, source)
                ws_corh.cell(i + 1, 12, competence)
                c_o = ws_corh.cell(i + 1, 15, amount_to_use)
                c_p = ws_corh.cell(i + 1, 16, current_balance - amount_to_use)
                c_o.number_format = c_p.number_format = "#,##0.00"
                ws_corh.cell(i, 17).value = None
                ws_corh.cell(i + 1, 17, "LINHA ACRESCENTADA +1 EMPENHO")
            else:
                ws_corh.cell(i, 8, note[11:])
                ws_corh.cell(i, 10, source)
                ws_corh.cell(i, 12, competence)
                c_o = ws_corh.cell(i, 15, amount_to_use)
                c_p = ws_corh.cell(i, 16, current_balance - amount_to_use)
                c_o.number_format = c_p.number_format = "#,##0.00"

            used_balances[note] = current_balance - amount_to_use
            remaining -= amount_to_use
            used_commitments += 1
            found = True

        if remaining > 0.01 and found:
            for col in range(1, 17):
                ws_corh.cell(i, col).fill = orange
            ws_log.cell(log_row, 1, i)
            ws_log.cell(log_row, 2, contract)
            ws_log.cell(log_row, 3, municipality)
            ws_log.cell(log_row, 4, remaining)
            ws_log.cell(log_row, 5, "Saldo insuficiente")
            log_row += 1

        i -= 1

    last_row_final = last_row(ws_corh)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    for row in range(5, last_row_final + 1):
        for col in range(1, 18):
            c = ws_corh.cell(row, col)
            c.border = border
            c.alignment = Alignment(horizontal=c.alignment.horizontal, vertical="center")
    for col in range(1, 18):
        c = ws_corh.cell(5, col)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
    for row in range(6, last_row_final + 1):
        ws_corh.cell(row, 15).number_format = "#,##0.00"
        ws_corh.cell(row, 16).number_format = "#,##0.00"
    autofit(ws_corh, 1, 17)
    autofit(ws_log, 1, 5)

    wb.save(path)

    print("Processamento concluido!\n")
    print(f"Linhas originais: {last_row_corh - 1}")
    print(f"Linhas inseridas: {inserted_rows}")
    print(f"Empenhos usados: {used_commitments}")
    print(f"Nao encontrados: {not_found}")
    print(f"Tempo: {time.time() - start_time:.2f} segundos")


if __name__ == "__main__":
    process_commitments()
