import openpyxl

from config import WORKBOOK_PATH

SHEET_CORH = "ajustada CORH"
MARKER = "LINHA ACRESCENTADA +1 EMPENHO"


def as_str(v):
    return str(v).strip() if v is not None else ""


def last_row(ws, col=1):
    lr = ws.max_row
    while lr > 1 and ws.cell(lr, col).value in (None, ""):
        lr -= 1
    return lr


def reset_commitments(path=WORKBOOK_PATH):
    wb = openpyxl.load_workbook(path, keep_vba=str(path).lower().endswith(".xlsm"))
    ws = wb[SHEET_CORH]

    lr = last_row(ws)
    deleted_rows = 0

    for i in range(lr, 1, -1):
        if as_str(ws.cell(i, 17).value) == MARKER:
            ws.delete_rows(i)
            deleted_rows += 1

    lr = last_row(ws)

    if lr >= 6:
        for row in range(6, lr + 1):
            ws.cell(row, 8).value = None
            ws.cell(row, 10).value = None
            ws.cell(row, 12).value = None

    ws.delete_cols(17)
    ws.delete_cols(16)
    ws.delete_cols(15)

    wb.save(path)

    print("Reset concluido!\n")
    print(f"Linhas removidas: {deleted_rows}")


if __name__ == "__main__":
    reset_commitments()
